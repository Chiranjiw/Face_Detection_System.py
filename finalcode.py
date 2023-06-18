# HELLO---------------------------------------------
import string
import tkinter as tk
from tkinter import messagebox
import csv

import cv2
import os
import numpy as np
from PIL import Image
import pandas as pd
import datetime
import time
import openpyxl
from openpyxl.reader.excel import load_workbook
import customtkinter
from CTkMessagebox import CTkMessagebox

customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")

root = customtkinter.CTk()
root.title('Face detection attendance system')
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
root.geometry(f"{screen_width}x{screen_height}+0+0")
frame = customtkinter.CTkFrame(master=root)
frame.pack(pady=50, padx=450, fill="both", expand=True)

dataSave = customtkinter.CTk()
dataSave.rowconfigure((0, 1, 2, 3, 4, 5), weight=1)
dataSave.columnconfigure(0, weight=1)
dataSave.minsize(200, 250)

def clear():
    std_Id.delete(0, 'end')  # Delete all the characters from the beginning (0) to end ('end')
    std_name.delete(0, 'end')
    std_roll.delete(0, 'end')
    std_age.delete(0, 'end')
    std_course.set('Select course')
    std_gender.set('Select gender')

class Admin:
    def createDatabase(self):
        Id = (std_Id.get())
        name = (std_name.get())
        without_spaces = name.replace(" ", "")
        roll = (std_roll.get())
        age = (std_age.get())
        gender = std_gender.get()
        course = std_course.get()

        if all(val != '' for val in [Id, without_spaces, roll, age, gender, course]):
            recCamera = cv2.VideoCapture(0)
            cascPath = os.path.dirname(cv2.__file__) + "/data/haarcascade_frontalface_default.xml"
            detector = cv2.CascadeClassifier(cascPath)
            frameCount = 0

            while True:
                ret, img = recCamera.read()
                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                faces = detector.detectMultiScale(gray, 1.1, 3)

                for (x, y, w, h) in faces:
                    cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
                    frameCount = frameCount + 1
                    # store each student picture with its name and roll
                    cv2.imwrite("E:\\ATTENDENCE_SYSTEM\\Students Faces\\ " + name + "." + Id + '.' + str(frameCount) + ".jpg", gray[y:y + h, x:x + h])
                    cv2.imshow('FACE RECOGNIZER', img)
                if cv2.waitKey(100) & 0xFF == ord('q' or 'Q'):
                    break
                # stop the camera when the number of picture exceed 50 pictures for each student
                if frameCount > 50:
                    break
            recCamera.release() #Close the camera
            cv2.destroyAllWindows() #Close the face capturing window

            # wb = openpyxl.load_workbook('E:\\ATTENDENCE_SYSTEM\\Student Databse.xlsx')
            column_names = ['ID', 'NAME', 'ROLL', 'AGE', 'COURSE', 'GENDER']



            # --------------------------------------
            # check if the file exists
            if os.path.isfile('E:\\ATTENDENCE_SYSTEM\\Student Database.xlsx'):
                # if the file exists, open it and append the new data to it
                wb = load_workbook(filename='E:\\ATTENDENCE_SYSTEM\\Student Database.xlsx')
                ws = wb.active
            else:
                # if the file does not exist, create a new file and add the new data to it
                wb = openpyxl.Workbook()
                ws = wb.active
                column_names = ['ID', 'NAME', 'ROLL', 'AGE', 'COURSE', 'GENDER']
                ws.append(column_names)

            # Data to store
            row = [Id, name, roll, age, course, gender]
            ws.append(row)
            wb.save('E:\\ATTENDENCE_SYSTEM\\Student Database.xlsx')
            # Train the images for face detection algorithm
            self.trainer()

            CTkMessagebox(title="Successful",message='Database created successfully: \nStudent Id: ' + Id + '\nStudent Name: ' + name, icon="check",option_1="OK")

        else:
            # root = tk.Tk()
            # root.geometry("100x100")
            # tk.messagebox.showwarning(title='WARNING!', message='All fields are compulsory')
            # root.mainloop()
            CTkMessagebox(title="WARNING!",
                          message='All fields are compulsory',
                          icon="warning", option_1="OK")

    def trainer(self):
        recognizer = cv2.face.LBPHFaceRecognizer_create()
        faces, Id = self.getImagesAndLabels("E:\\ATTENDENCE_SYSTEM\\Students Faces")
        recognizer.train(faces, np.array(Id))
        recognizer.save("E:\\ATTENDENCE_SYSTEM\\Image Trainer.yml")

    def getImagesAndLabels(self, path):
            imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
            faces = []
            Ids = []
            for imagePath in imagePaths:
                pilImage = Image.open(imagePath).convert('L')
                imageNp = np.array(pilImage, 'uint8')
                Id = int(os.path.split(imagePath)[-1].split(".")[1])
                faces.append(imageNp)
                Ids.append(Id)
            return faces, Ids

class Faculty:
    def takeAttendance(self):
        recognizer = cv2.face.LBPHFaceRecognizer_create()
        recognizer.read("E:\\ATTENDENCE_SYSTEM\\Image Trainer.yml")
        faceCascadePath = os.path.dirname(cv2.__file__) + "/data/haarcascade_frontalface_default.xml"
        faceCascade = cv2.CascadeClassifier(faceCascadePath)

        eyeCascadePath = os.path.dirname(cv2.__file__) + "/data/haarcascade_eye.xml"
        eyeCascade = cv2.CascadeClassifier(eyeCascadePath)

        df = pd.read_excel("E:\\ATTENDENCE_SYSTEM\\Student Database.xlsx")
        font = cv2.FONT_HERSHEY_COMPLEX_SMALL
        attCamera = cv2.VideoCapture(0)

        # create a column to hold the student id,name,date and time
        row2 = ['Id', 'Name', 'Time']
        attendance = pd.DataFrame(columns=row2)

        # attdColumn = ['ID', 'NAME', 'TIME']
        ts = time.time()
        currDate = datetime.datetime.fromtimestamp(ts).strftime("%Y-%m-%d")

        # create or load the workbook
        try:
            workbook = openpyxl.load_workbook('E:\\ATTENDENCE_SYSTEM\\' + currDate + '.xlsx')
            worksheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.append(row2)

        while True:
            ret, img = attCamera.read()
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = faceCascade.detectMultiScale(gray, 1.3, 5)
            for (x, y, w, h) in faces:
                cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
                # ------------Eye Detection-----------
                # cropping face only
                roi_color = img[y:y + h, x:x + w]
                roi_gray = gray[y:y + h, x:x + w]
                # eyes_detection
                eyes = eyeCascade.detectMultiScale(roi_gray)
                # Eyes rectangles
                for (ex, ey, ew, eh) in eyes:
                    cv2.rectangle(roi_color, (ex, ey), (ex + ew, ey + eh), (255, 255, 0), 2)
                # ------------------------------------
                Id, conf = recognizer.predict(gray[y:y + h, x:x + w])

                # A confidence less than 60 indicates a good face recognition
                if conf < 60:
                    ts = time.time()
                    timeStamp = datetime.datetime.fromtimestamp(ts).strftime('%H:%M:%S')
                    name = df.loc[df["ID"] == Id]["NAME"].values
                    rmvName = str(name).strip("[]'")
                    display_name = str(Id) + "-" + name

                    if Id and len(attendance[attendance['Id'] == Id]) == 0:
                        row2 = attendance.loc[len(attendance)] = [Id, rmvName, timeStamp]
                        worksheet.append(row2)
                else:
                    display_name = "UNKNOWN"

                # show the student id and name
                cv2.putText(img, str(display_name), (x, y + h - 10), font, 0.8, (255, 255, 255), 1)
                cv2.imshow('PRESS Q TO CLOSE', img)
            if cv2.waitKey(1) & 0xFF == ord('q' or 'Q'):
                break
        # Save the attendance details
        workbook.save('E:\\ATTENDENCE_SYSTEM\\' + currDate + '.xlsx')

        CTkMessagebox(title="ATTENDANCE",
                      message='Attendance updated successfully',
                      icon="check", option_1="OK")

        # Close camera
        attCamera.release()
        # Close all open windows
        cv2.destroyAllWindows()


label = customtkinter.CTkLabel(master=frame, text= "FACE ATTENDANCE SYSTEM", font=("Roboto",24))
label.pack(pady=50, padx=10)

# Id ---------------------------------------------
def check_entry(id):
    if id.isdigit() or id == "":
        return True
    else:

        CTkMessagebox(title="Error",
                      message='Please enter a valid integer',
                      icon="cancel", option_1="OK")
        return False


std_Id = customtkinter.CTkEntry(master=frame, placeholder_text="Id", width=300)
std_Id.pack(pady=12, padx=10)

std_name = customtkinter.CTkEntry(master=frame, placeholder_text="Student name",width=300)
std_name.pack(pady=12, padx=10)


std_roll = customtkinter.CTkEntry(master=frame, placeholder_text="Student roll",width=300)
std_roll.pack(pady=12, padx=10)


std_age = customtkinter.CTkEntry(master=frame, placeholder_text="Student age",width=300)
std_age.pack(pady=12, padx=10)


std_course= customtkinter.StringVar()
comboBox = customtkinter.CTkComboBox(master=frame, values=['BCA','MCA','BBA','MBA'],state= 'readonly',variable=std_course,width=300)
comboBox.set('Select course')
comboBox.pack(pady=12, padx=10)


std_gender= customtkinter.StringVar()
comboBox = customtkinter.CTkComboBox(master=frame, values=['Male','Female'],state= 'readonly',variable=std_gender,width=300)
comboBox.set('Select gender')
comboBox.pack(pady=12, padx=10)




clearBtn1 = customtkinter.CTkButton(master=frame, text="Clear all", command=clear,width=200, fg_color="#7E3517", hover_color="maroon")
clearBtn1.pack(pady=12, padx=10)


takeImageBtn = customtkinter.CTkButton(master=frame, text="Create Database", command=lambda: Admin().createDatabase(),width=200)
takeImageBtn.pack(pady=12, padx=10)



trackImageBtn = customtkinter.CTkButton(master=frame, text="Take Attendance", command=lambda: Faculty().takeAttendance(),width=200)
trackImageBtn.pack(pady=12, padx=10)

#-----------------------------------

root.mainloop()

